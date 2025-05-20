import re
import numpy as np
import statistics
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os

archivo = "resultado_terremotos.txt"

# Expresiones regulares 
patron_consulta = re.compile(r"Consulta para '(.+?)' del (\d{4}-\d{2}-\d{2}) al (\d{4}-\d{2}-\d{2}):")
patron_magnitud = re.compile(r"- Magnitud ([\d.]+): (\d+)")

# Lista estructurada para todos los datos
datos_estructurados = []

# Variables temporales
region_actual = None
fecha_inicio = None
fecha_fin = None
magnitudes_actuales = []

# Leer archivo y estructurar datos
with open(archivo, "r", encoding="utf-8") as f:
    for linea in f:
        linea = linea.strip()

        match_consulta = patron_consulta.match(linea)
        if match_consulta:
            if region_actual and magnitudes_actuales:
                try:
                    moda = statistics.mode(magnitudes_actuales)
                except statistics.StatisticsError:
                    moda = None

                datos_estructurados.append({
                    "region": region_actual,
                    "fecha_inicio": fecha_inicio,
                    "fecha_fin": fecha_fin,
                    "magnitudes": magnitudes_actuales,
                    "frecuencia": dict(sorted(
                        {m: magnitudes_actuales.count(m) for m in set(magnitudes_actuales)}.items())),
                    "media": round(np.mean(magnitudes_actuales), 2),
                    "moda": moda
                })

            region_actual = match_consulta.group(1).lower()
            fecha_inicio = match_consulta.group(2)
            fecha_fin = match_consulta.group(3)
            magnitudes_actuales = []
            continue

        match_magnitud = patron_magnitud.match(linea)
        if match_magnitud:
            magnitud = float(match_magnitud.group(1))
            cantidad = int(match_magnitud.group(2))
            magnitudes_actuales.extend([magnitud] * cantidad)

# Guardar última región
if region_actual and magnitudes_actuales:
    try:
        moda = statistics.mode(magnitudes_actuales)
    except statistics.StatisticsError:
        moda = None

    datos_estructurados.append({
        "region": region_actual,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "magnitudes": magnitudes_actuales,
        "frecuencia": dict(sorted(
            {m: magnitudes_actuales.count(m) for m in set(magnitudes_actuales)}.items())),
        "media": round(np.mean(magnitudes_actuales), 2),
        "moda": moda
    })

# GRAFICAS

def graficar_region(entrada):
    region = entrada["region"]
    frecuencia = entrada["frecuencia"]
    magnitudes = list(frecuencia.keys())
    cantidades = list(frecuencia.values())

    if not magnitudes or not cantidades or len(magnitudes) != len(cantidades):
        print(f"No se puede graficar {region}: datos incompletos.")
        return

    # Linea
    plt.figure(figsize=(8, 5))
    plt.plot(magnitudes, cantidades, marker='o', linestyle='-', color='b')
    plt.title(f"Frecuencia de Magnitudes en {region.title()}")
    plt.xlabel("Magnitud")
    plt.ylabel("Frecuencia")
    plt.grid(True)
    plt.savefig(f"grafica_lineas_{region}.png")
    plt.close()

    # Barras
    plt.figure(figsize=(8, 5))
    plt.bar([str(m) for m in magnitudes], cantidades, color='orange')
    plt.title(f"Distribución de Magnitudes - {region.title()}")
    plt.xlabel("Magnitud")
    plt.ylabel("Frecuencia")
    plt.grid(axis='y')
    plt.savefig(f"grafica_barras_{region}.png")
    plt.close()

    # Dispersión
    plt.figure(figsize=(8, 5))
    dispersos = [(i+1, mag) for i, mag in enumerate(entrada["magnitudes"])]
    x, y = zip(*dispersos)
    plt.scatter(x, y, color='green', alpha=0.6)
    plt.title(f"Dispersión de Magnitudes - {region.title()}")
    plt.xlabel("Evento (orden cronológico)")
    plt.ylabel("Magnitud")
    plt.grid(True)
    plt.savefig(f"grafica_dispersion_{region}.png")
    plt.close()

    # Pastel
    plt.figure(figsize=(7, 7))
    plt.pie(cantidades, labels=[str(m) for m in magnitudes], autopct='%1.1f%%', startangle=140)
    plt.title(f"Proporción de Magnitudes - {region.title()}")
    plt.axis('equal')
    plt.savefig(f"grafica_pastel_{region}.png")
    plt.close()

# Crear libro de Excel
libro = Workbook()
pagina = libro.active
pagina.title = "Datos obtenidos"
pagina.append(["Región", "Fecha Inicio", "Fecha Fin", "Media", "Moda"])

# Recorrer y graficar cada entrada
for entrada in datos_estructurados:
    region = entrada["region"]

    # Generar gráficas
    graficar_region(entrada)

    # Agregar resumen a hoja principal
    pagina.append([
        entrada["region"],
        entrada["fecha_inicio"],
        entrada["fecha_fin"],
        entrada["media"],
        entrada["moda"]
    ])

    # Crear hoja individual
    nombre_hoja = region[:31]
    hoja = libro.create_sheet(title=nombre_hoja)
    hoja.append(["Magnitud", "Frecuencia"])
    for mag, freq in entrada["frecuencia"].items():
        hoja.append([mag, freq])

    # Insertar imágenes
    graficas = [
        f"grafica_lineas_{region}.png",
        f"grafica_barras_{region}.png",
        f"grafica_dispersion_{region}.png",
        f"grafica_pastel_{region}.png"
    ]

    fila_imagen = len(entrada["frecuencia"]) + 4
    for grafica in graficas:
        if os.path.exists(grafica):
            img = ExcelImage(grafica)
            img.width = 480
            img.height = 280
            hoja.add_image(img, f"A{fila_imagen}")
            fila_imagen += 22  

# Guardar archivo final
libro.save("Datos_Terremotos.xlsx")
print("Archivo Excel generado con gráficas incluidas.")
